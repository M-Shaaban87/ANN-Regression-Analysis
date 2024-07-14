'''
This script effectively creates, trains, and evaluates multiple ANN models, 
for regression analysis and prediction.
Created by: Dr. Mohammed Shaaban ... Assistant Professor (Structural Engineering)
Email: mohamed.selim@deltauniv.edu.eg
'''
import pandas as pd
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
import tensorflow as tf
from tensorflow.keras.models import Sequential, load_model
from tensorflow.keras.layers import Dense, Dropout
from tensorflow.keras.callbacks import EarlyStopping
import matplotlib.pyplot as plt
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the data
data = pd.read_excel('data.xlsx')

# Assuming the 'H UAV' column is the target and the rest are features
X = data.iloc[:, :-1].values  # Features (all columns except the last one)
y = data.iloc[:, -1].values   # Target (the last column)

# Split the data into training and test sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Scale the features
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)

# Define different ANN architectures
def create_ann_model(input_shape, dense_units, dropout_rate=0.0):
    model = Sequential()
    model.add(Dense(dense_units[0], input_dim=input_shape, activation='relu'))
    for units in dense_units[1:]:
        model.add(Dense(units, activation='relu'))
        if dropout_rate > 0:
            model.add(Dropout(dropout_rate))
    model.add(Dense(1))
    model.compile(optimizer='adam', loss='mse')
    return model

# Create 20 different ANN architectures
models = [
    ([50], 0.0),
    ([100], 0.0),
    ([50, 50], 0.0),
    ([100, 50], 0.0),
    ([50, 50, 50], 0.0),
    ([100, 50, 50], 0.0),
    ([50, 50, 50], 0.2),
    ([100, 50, 50], 0.2),
    ([50, 100, 50], 0.2),
    ([100, 100, 50], 0.2),
    ([50, 50], 0.2),
    ([100, 50], 0.2),
    ([50, 50, 50], 0.2),
    ([100, 50, 50], 0.2),
    ([50, 100, 50], 0.2),
    ([100, 100, 50], 0.2),
    ([50, 50, 50], 0.5),
    ([100, 50, 50], 0.5),
    ([50, 100, 50], 0.5),
    ([100, 100, 50], 0.5),
]

# Train and evaluate each model
best_model = None
best_mse = float('inf')
history_dict = {}
results = []

for i, (dense_units, dropout_rate) in enumerate(models):
    model = create_ann_model(X_train.shape[1], dense_units, dropout_rate)
    early_stopping = EarlyStopping(monitor='val_loss', patience=5, restore_best_weights=True)
    history = model.fit(X_train, y_train, validation_split=0.2, epochs=50, callbacks=[early_stopping], verbose=0)
    
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    r2 = r2_score(y_test, y_pred)
    mae = mean_absolute_error(y_test, y_pred)
    results.append((i + 1, dense_units, dropout_rate, mse, r2, mae))
    
    if mse < best_mse:
        best_mse = mse
        best_model = model
        best_model_details = (dense_units, dropout_rate)
        best_y_pred = y_pred.flatten()  # Flatten to ensure it matches the shape of y_test
        best_r2 = r2
        
    history_dict[f'Model_{i + 1}'] = history

# Create output directories
os.makedirs('Charts', exist_ok=True)
os.makedirs('Tables', exist_ok=True)

# Save the best model
best_model.save('best_ann_model.keras')

# Convert results to DataFrame
results_df = pd.DataFrame(results, columns=['Model', 'Dense Units', 'Dropout Rate', 'MSE', 'R^2', 'MAE'])
results_df.to_excel('Tables/model_comparison.xlsx', index=False)

# Highlight the best model in the Excel file
wb = load_workbook('Tables/model_comparison.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=2, max_col=6, max_row=21):
    if row[3].value == best_mse:
        for cell in row:
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
wb.save('Tables/model_comparison.xlsx')

# Plot training and validation loss for each model
for model_name, history in history_dict.items():
    plt.figure()
    plt.plot(history.history['loss'], label='Training Loss')
    plt.plot(history.history['val_loss'], label='Validation Loss')
    plt.title(model_name)
    plt.xlabel('Epochs')
    plt.ylabel('Loss')
    plt.legend()
    plt.savefig(f'Charts/{model_name}_loss.png')
    plt.close()

# Plot comparison of MSE
plt.figure()
plt.bar(results_df['Model'], results_df['MSE'])
plt.xlabel('Model')
plt.ylabel('Mean Squared Error')
plt.title('Comparison of MSE for Different ANN Models')
plt.savefig('Charts/comparison_mse.png')
plt.close()

# Plot comparison of R^2
plt.figure()
plt.bar(results_df['Model'], results_df['R^2'])
plt.xlabel('Model')
plt.ylabel('R^2 Score')
plt.title('Comparison of R^2 for Different ANN Models')
plt.savefig('Charts/comparison_r2.png')
plt.close()

# Plot comparison of MAE
plt.figure()
plt.bar(results_df['Model'], results_df['MAE'])
plt.xlabel('Model')
plt.ylabel('Mean Absolute Error')
plt.title('Comparison of MAE for Different ANN Models')
plt.savefig('Charts/comparison_mae.png')
plt.close()

# Plot residuals
plt.figure()
plt.scatter(y_test, y_test - best_y_pred)
plt.xlabel('True Values')
plt.ylabel('Residuals')
plt.title('Residuals of Best Model')
plt.axhline(y=0, color='r', linestyle='--')
plt.savefig('Charts/residuals.png')
plt.close()

# Plot true vs predicted values
plt.figure()
plt.scatter(y_test, best_y_pred)
plt.plot([min(y_test), max(y_test)], [min(y_test), max(y_test)], color='r', linestyle='--')
plt.xlabel('True Values')
plt.ylabel('Predicted Values')
plt.title(f'True vs Predicted Values (R^2 = {best_r2:.2f})')
plt.savefig('Charts/true_vs_predicted.png')
plt.close()

# Save weights and biases of the best model with more details
weights = best_model.get_weights()
layer_names = [layer.name for layer in best_model.layers]
weight_shapes = [weight.shape for weight in weights]

details = []
for layer_name, weight_shape, weight in zip(layer_names, weight_shapes, weights):
    details.append({
        'Layer Name': layer_name,
        'Weight Shape': weight_shape,
        'Weights': weight.flatten().tolist()  # Convert to list for better readability
    })

# Convert the details to a DataFrame and save to Excel
weights_biases_df = pd.DataFrame(details)
weights_biases_df.to_excel('Tables/best_model_weights_biases.xlsx', index=False)

# Print details of the best ANN
print("Best ANN Model Details:")
print(f"Dense Units: {best_model_details[0]}")
print(f"Dropout Rate: {best_model_details[1]}")
print(f"MSE: {best_mse}")
print(f"R^2: {best_r2}")
